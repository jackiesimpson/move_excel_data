import openpyxl
OE = "E:\\haha1.xlsx"
DE = "E:\\haha2.xlsx"
def readFromOE(path):
    book = openpyxl.load_workbook(path)
    sheet = book.worksheets[0]
    row = []
    col = []
    for cell in list(sheet.rows)[0]:
        row.append(cell.value)
    for cell in list(sheet.columns)[0]:
        col.append(cell.value)
    #print(sheet.cell(1,2).value)
    return row,col

def moveData(seq_Orow,seq_Ocol,seq_Drow,seq_Dcol):
    book1 = openpyxl.load_workbook(OE)
    book2 = openpyxl.load_workbook(DE)
    sheet1 = book1.worksheets[0]
    sheet2 = book2.worksheets[0]
    for i in range(0,len(seq_Orow)):
        for j in range(0,len(seq_Ocol)):
            #print(seq_Orow[i],seq_Ocol[j])
            #print(seq_Drow[i],seq_Dcol[j])
            sheet2.cell(seq_Drow[i],seq_Dcol[j],sheet1.cell(seq_Orow[i],seq_Ocol[j]).value)
    book2.save(DE)
def get_seq(num,startpoint):
    seq = []
    for i in range(startpoint,num+1):
        seq.append(i)
    return seq


if __name__ =="__main__":
    Orow,Ocol = readFromOE(OE)
    Drow,Dcol = readFromOE(DE)
    rowseq = get_seq(len(Orow),2)
    colseq = get_seq(len(Ocol),2)
    dic_Orow = dict(zip(Orow[1:],rowseq))
    dic_Ocol = dict(zip(Ocol[1:],colseq))

    dic_Drow = dict(zip(Drow[1:],rowseq))
    dic_Dcol = dict(zip(Dcol[1:],colseq))

    # print(dic_Orow)
    # print(dic_Ocol)
    # print(dic_Drow)
    # print(dic_Dcol)
    newseq_row = []
    newseq_col = []
    for i in dic_Orow.keys():
        newseq_row.append(dic_Drow[i])
    for i in dic_Ocol.keys():
        newseq_col.append(dic_Dcol[i])
    moveData(rowseq,colseq,newseq_col,newseq_row)
